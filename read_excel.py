import pandas as pd
import json
import sys
from openpyxl import load_workbook

file_path = r"C:/Users/tempv2/Desktop/PortfolioAgent/Ref Docs/Work/Marketing/Testimonial Creation/All Testimonial Creation Resources.xlsx"

try:
    # Load workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Find URL column index
    try:
        url_col_idx = headers.index("URL")
    except ValueError:
        print("URL column not found")
        sys.exit(1)
        
    data = []
    
    # Iterate rows (skip header)
    for row in ws.iter_rows(min_row=2):
        row_data = {}
        for idx, cell in enumerate(row):
            header = headers[idx]
            
            # If this is the URL column, check for hyperlink
            if idx == url_col_idx and cell.hyperlink:
                value = cell.hyperlink.target
            else:
                value = cell.value
                
            row_data[header] = value
            
        # Filter empty rows
        if row_data.get("Resource"):
            data.append(row_data)
    
    # Print as JSON
    print(json.dumps(data, indent=2, default=str))

except ImportError:
    print("Error: openpyxl not installed.")
except Exception as e:
    print(f"Error reading file: {e}")
