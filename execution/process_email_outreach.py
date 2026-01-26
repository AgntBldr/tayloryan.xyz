import pandas as pd
import json
import os
import glob
import openpyxl

# Configuration
BASE_DIR = r"C:\Users\tempv2\Desktop\PortfolioAgent"
INPUT_DIR = r"C:\Users\tempv2\Desktop\PortfolioAgent\Ref Docs\Work\Email Outreach"
OUTPUT_FILE = os.path.join(BASE_DIR, "assets", "js", "email_outreach_data.js")

def normalize_column_name(col):
    return str(col).strip().lower().replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "")

def get_col_index(columns, possible_names):
    for i, col in enumerate(columns):
        if normalize_column_name(col) in possible_names:
            return i + 1 # openpyxl is 1-indexed
    return None

def process_file_with_links(filepath, resource_type_default="Resource"):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False) # data_only=False to get hyperlinks? Actually hyperlinks are available anyway.
        all_data = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Read header row (assuming row 1)
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            if not headers:
                continue

            normalized_headers = [normalize_column_name(h) for h in headers]
            # print(f"Processing {os.path.basename(filepath)} - sheet: {sheet_name}. Headers: {normalized_headers}")

            # Identify Column Indices
            title_idx = get_col_index(headers, ['name', 'title', 'resource_name', 'template_name', 'subject_line', 'resource', 'prompt_name'])
            # Fallback for title: First object-like column? Harder in openpyxl. Let's rely on specific names or first column.
            if not title_idx: title_idx = 1 

            link_idx = get_col_index(headers, ['url', 'link', 'website', 'resource_link', 'hyperlink'])
            # If no explicit link column, check if Title column has hyperlinks? 
            
            desc_idx = get_col_index(headers, ['description', 'notes', 'summary', 'body', 'content'])
            type_idx = get_col_index(headers, ['type', 'category', 'format'])
            tags_idx = get_col_index(headers, ['tags', 'keywords', 'topic'])

            # Iterate rows (starting from 2)
            for row in ws.iter_rows(min_row=2):
                # Title
                title_cell = row[title_idx - 1] if title_idx and len(row) >= title_idx else None
                title = title_cell.value if title_cell else None
                
                if not title:
                    continue

                # URL logic:
                # 1. Check if 'link' column exists and has hyperlink or value
                # 2. Check if 'title' column has hyperlink
                url = ""
                
                if link_idx and len(row) >= link_idx:
                    link_cell = row[link_idx - 1]
                    if link_cell.hyperlink:
                        url = link_cell.hyperlink.target
                    elif link_cell.value:
                        url = str(link_cell.value)
                
                # Fallback: Check title cell for hyperlink
                if not url and title_cell.hyperlink:
                    url = title_cell.hyperlink.target
                
                # Cleaning URL
                if url and not str(url).lower().startswith(('http', 'mailto', 'docs.google.com')):
                     # Sometimes internal links or file paths? Ignore for now if junk.
                     pass

                # Description
                desc = ""
                if desc_idx and len(row) >= desc_idx:
                    desc = row[desc_idx - 1].value or ""

                # Type
                item_type = resource_type_default
                if type_idx and len(row) >= type_idx:
                    item_type = row[type_idx - 1].value or resource_type_default

                # Tags
                tags = ""
                if tags_idx and len(row) >= tags_idx:
                    tags = row[tags_idx - 1].value or ""

                # Create ID
                item_id = str(title).strip().lower().replace(" ", "-")
                import re
                item_id = re.sub(r'[^a-zA-Z0-9-]', '', item_id)[:50]

                item = {
                    "id": item_id,
                    "title": str(title).strip(),
                    "description": str(desc).strip(),
                    "url": str(url).strip(),
                    "type": str(item_type).strip(),
                    "tags": [t.strip() for t in str(tags).split(',')] if tags else [],
                    "source_file": os.path.basename(filepath),
                    "sheet_name": sheet_name
                }
                all_data.append(item)
                
        return all_data

    except Exception as e:
        print(f"Error processing {filepath}: {e}")
        return []

def main():
    all_resources = []
    
    # 1. Process "All Email Outreach Resources"
    resource_files = glob.glob(os.path.join(INPUT_DIR, "All Email*"))
    for f in resource_files:
        print(f"Reading Resources file: {f}")
        data = process_file_with_links(f, resource_type_default="Resource")
        all_resources.extend(data)

    # 2. Process "Email Outreach Templates"
    template_files = glob.glob(os.path.join(INPUT_DIR, "Email Outreach Templates*"))
    for f in template_files:
        print(f"Reading Templates file: {f}")
        data = process_file_with_links(f, resource_type_default="Template")
        all_resources.extend(data)

    print(f"Total extracted items: {len(all_resources)}")
    
    # Write to JS
    js_content = f"window.EMAIL_OUTREACH_DATA = {json.dumps(all_resources, indent=2)};"
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(js_content)
    
    print(f"Saved data to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
