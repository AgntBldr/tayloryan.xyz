import os
import json
import pandas as pd
# import docx # Will use if available, or text extraction fallback

# Configuration
BASE_DIR = r"C:\Users\tempv2\Desktop\PortfolioAgent"
RESOURCES_DIR = os.path.join(BASE_DIR, "Ref Docs", "Work", "Quests", "Quest Resources")
OUTPUT_FILE = os.path.join(BASE_DIR, "assets", "resources_data.js")

CATEGORIES = {
    "Discovery Resources": "Project Discovery",
    "Enrichment Resources": "Quest Enrichment",
    "Quest Building": "Quest Building",
    "Quest Promotion": "Quest Promotion"
}

def extract_excel(filepath):
    try:
        # Read all sheets
        xls = pd.ExcelFile(filepath)
        sheets_data = {}
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            # Replace NaN with empty string/null for JSON
            df = df.fillna("")
            sheets_data[sheet_name] = df.to_dict(orient='records')
        return {"type": "Sheet", "content": sheets_data}
    except Exception as e:
        print(f"Error reading Excel {filepath}: {e}")
        return None

def extract_docx(filepath):
    try:
        import docx
        doc = docx.Document(filepath)
        html_parts = []
        list_state = None # 'ul', 'ol', or None
        
        import re
        def linkify_text(text):
            return re.sub(r'(https?://\S+)', r'<a href="\1" target="_blank" class="text-blue-400 hover:underline">\1</a>', text)

        def close_list():
            nonlocal list_state
            if list_state:
                html_parts.append(f"</{list_state}>")
                list_state = None

        for para in doc.paragraphs:
            raw_text = para.text.strip()
            if not raw_text:
                continue
            
            text = linkify_text(raw_text)
            style_name = para.style.name.lower()
            
            # Headings
            if 'heading' in style_name:
                close_list()
                level = 3
                if '1' in style_name: level = 3
                elif '2' in style_name: level = 4
                elif '3' in style_name: level = 5
                html_parts.append(f"<h{level} class='text-white font-bold mb-2 mt-4'>{text}</h{level}>")
            
            # Lists
            elif 'list' in style_name or 'bullet' in style_name:
                wanted_state = 'ol' if 'number' in style_name else 'ul'
                if list_state != wanted_state:
                    close_list()
                    list_state = wanted_state
                    html_parts.append(f"<{list_state} class='list-disc pl-5 mb-4 space-y-1 text-neutral-300'>")
                
                html_parts.append(f"<li>{text}</li>")
            
            # Normal text
            else:
                close_list()
                if raw_text.startswith('- ') or raw_text.startswith('• '):
                     if list_state != 'ul':
                        list_state = 'ul'
                        html_parts.append(f"<ul class='list-disc pl-5 mb-4 space-y-1 text-neutral-300'>")
                     # Remove bullet char but keep linkified text (careful with indexing if link is at start)
                     # safer to linkify AFTER removing bullet
                     clean_text = raw_text[2:]
                     html_parts.append(f"<li>{linkify_text(clean_text)}</li>")
                else:
                    html_parts.append(f"<p class='mb-4 text-neutral-300 leading-relaxed'>{text}</p>")

        close_list() # Close any remaining list
        
        return {"type": "Doc", "content": "\n".join(html_parts)}

    except ImportError:
        return {"type": "Doc", "content": "<p class='text-red-400'>Error: python-docx not installed.</p>"}
    except Exception as e:
        print(f"Error reading Docx {filepath}: {e}")
        return {"type": "Doc", "content": f"<p class='text-red-400'>Error reading file: {e}</p>"}

# ... (imports remain)

METADATA_FILE = os.path.join(RESOURCES_DIR, "All Quest Portfolio Resources.xlsx")

def load_metadata():
    """Reads the master Excel file and returns a lookup dict using openpyxl to capture hyperlinks."""
    if not os.path.exists(METADATA_FILE):
        print(f"Metadata file not found: {METADATA_FILE}")
        return {}
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(METADATA_FILE)
        # Assume first sheet is the one we want
        ws = wb.worksheets[0]
        
        # Find column indices (0-based)
        headers = [cell.value for cell in ws[1]]
        headers = [str(h).strip().lower() for h in headers]
        
        idx_resource = -1
        idx_url = -1
        idx_desc = -1
        
        # Heuristic for columns
        for i, h in enumerate(headers):
            if h in ['resource', 'name', 'title', 'project']: idx_resource = i
            elif h in ['url', 'link', 'website']: idx_url = i
            elif h in ['description', 'desc']: idx_desc = i
            
        if idx_resource == -1:
            print("Could not find Resource/Title column in metadata")
            return {}

        metadata = {}
        # Iterate rows starting from 2
        for row in ws.iter_rows(min_row=2):
            title_cell = row[idx_resource]
            title = title_cell.value
            
            if not title: continue
            
            desc_val = ""
            if idx_desc != -1:
                desc_val = row[idx_desc].value
            
            url_val = ""
            if idx_url != -1:
                url_cell = row[idx_url]
                if url_cell.hyperlink:
                    url_val = url_cell.hyperlink.target
                elif url_cell.value:
                    url_val = str(url_cell.value)
            
            # Key normalization
            key = str(title).replace("[Sheet]", "").replace("[Doc]", "").strip().lower()
            metadata[key] = {
                "description": str(desc_val) if desc_val else "",
                "external_url": str(url_val) if url_val else ""
            }
            
        return metadata

    except Exception as e:
        print(f"Error reading metadata: {e}")
        return {}

def main():
    resources_db = {}
    metadata_lookup = load_metadata()
    print(f"Loaded {len(metadata_lookup)} metadata entries.")

    for folder_name, display_category in CATEGORIES.items():
        folder_path = os.path.join(RESOURCES_DIR, folder_name)
        if not os.path.exists(folder_path):
            print(f"Warning: {folder_path} not found.")
            continue
            
        resources_db[display_category] = []
        
        for filename in os.listdir(folder_path):
            filepath = os.path.join(folder_path, filename)
            if filename.startswith("~$"): continue 
            
            # Basic info
            title = os.path.splitext(filename)[0]
            resource_item = {
                "title": title,
                "filename": filename,
                "description": "",
                "external_url": ""
            }
            
            # Metadata Merge
            lookup_key = title.strip().lower()
            if lookup_key in metadata_lookup:
                meta = metadata_lookup[lookup_key]
                resource_item["description"] = meta["description"]
                resource_item["external_url"] = meta["external_url"]
            
            # Content Extraction
            if filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv"):
                if filename.endswith(".csv"):
                     try:
                        df = pd.read_csv(filepath)
                        df = df.fillna("")
                        resource_item.update({"type": "Sheet", "content": {"Sheet1": df.to_dict(orient='records')}})
                     except: pass
                else:
                    data = extract_excel(filepath)
                    if data: resource_item.update(data)
            
            elif filename.endswith(".docx"):
                data = extract_docx(filepath)
                if data: resource_item.update(data)
            
            else:
                continue

            resources_db[display_category].append(resource_item)

    # Wrap in JS variable
    js_content = f"const RESOURCES_DB = {json.dumps(resources_db, indent=2)};"
    
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(js_content)
    
    print(f"Successfully wrote resources db to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
