import pandas as pd
import json
import os
import glob
import openpyxl
import re

# Configuration
BASE_DIR = r"C:\Users\tempv2\Desktop\PortfolioAgent"
MARKETING_DIR = os.path.join(BASE_DIR, "Ref Docs", "Work", "Marketing")
OUTPUT_FILE = os.path.join(BASE_DIR, "assets", "js", "marketing_full_data.js")

CATEGORIES = {
    "Content Creator Program": "Content Creator",
    "Email Outreach": "Email Outreach",
    "Affiliate Program": "Affiliates",
    "Case Study Creation": "Case Studies"
}

def normalize_column_name(col):
    return str(col).strip().lower().replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "").replace("-", "_")

def get_col_index(columns, possible_names):
    for i, col in enumerate(columns):
        if normalize_column_name(col) in possible_names:
            return i + 1 # openpyxl is 1-indexed
    return None

def process_file_with_links(filepath, category):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        all_data = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Read header row (assuming row 1)
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            if not headers:
                continue

            # Identify Column Indices
            title_idx = get_col_index(headers, ['name', 'title', 'resource_name', 'template_name', 'subject_line', 'resource', 'prompt_name', 'case_study', 'company', 'service'])
            if not title_idx: title_idx = 1 # Fallback to first column

            link_idx = get_col_index(headers, ['url', 'link', 'website', 'resource_link', 'hyperlink', 'example_link', 'drive_link'])
            
            desc_idx = get_col_index(headers, ['description', 'notes', 'summary', 'body', 'content', 'purpose', 'usage'])
            type_idx = get_col_index(headers, ['type', 'category', 'format', 'asset_type'])
            tags_idx = get_col_index(headers, ['tags', 'keywords', 'topic', 'niche'])
            
            # Additional Fields for specific types
            status_idx = get_col_index(headers, ['status', 'state'])
            owner_idx = get_col_index(headers, ['owner', 'creator'])

            # Iterate rows (starting from 2)
            current_section = "General"
            
            for row in ws.iter_rows(min_row=2):
                # Title
                title_cell = row[title_idx - 1] if title_idx and len(row) >= title_idx else None
                title = title_cell.value if title_cell else None
                
                if not title:
                    continue

                # User-defined Exclusions
                EXCLUDED_TITLES = [
                    "Links to Your Posts", "Links to Comments", 
                    "Monthly Posts", "Tag Us", "Playbook / Case Study", "Report (Email)",
                    "Name", "Linkedin", "Email", "Online Calendar",
                    "Website [URL]", "Platform [URL]", "Api Documentation [URL]",
                    "Affiliate / Rferral Registration", "Affiliate / Referral Registration"
                ]
                if str(title).strip() in EXCLUDED_TITLES:
                    continue

                # URL Logic
                url = ""
                if link_idx and len(row) >= link_idx:
                    link_cell = row[link_idx - 1]
                    if link_cell.hyperlink:
                        url = link_cell.hyperlink.target
                    elif link_cell.value:
                        url = str(link_cell.value)
                
                # Fallback: Check title cell for hyperlink
                if not url and title_cell and title_cell.hyperlink:
                    url = title_cell.hyperlink.target

                # Description
                desc = ""
                if desc_idx and len(row) >= desc_idx:
                    desc = row[desc_idx - 1].value or ""

                # Type
                item_type = "Resource"
                if type_idx and len(row) >= type_idx:
                    item_type = row[type_idx - 1].value or "Resource"

                # Check if this is a "Blue Header" / Section Separator
                # Heuristic: No URL, No Description (or very short), and Title exists
                # Also check standard headers user mentioned
                is_header = False
                if not url and (not desc or len(str(desc)) < 5):
                     is_header = True
                
                # If specific known headers, force it
                known_headers = ["strategy", "admin tools", "targeting & discovery", "outreach", "crm tools", "scheduling tools directory"]
                if str(title).lower().strip() in known_headers:
                    is_header = True

                if is_header:
                    current_section = str(title).strip()
                    continue # Skip adding this as an item

                # STRICT URL REQUIREMENT (User Request)
                if not url:
                    continue

                # Tags
                tags = []
                if tags_idx and len(row) >= tags_idx:
                    val = row[tags_idx - 1].value
                    if val:
                        tags = [t.strip() for t in str(val).split(',')]

                # Status & Owner
                status = ""
                if status_idx and len(row) >= status_idx:
                    status = row[status_idx - 1].value
                
                owner = ""
                if owner_idx and len(row) >= owner_idx:
                    owner = row[owner_idx - 1].value

                # Enhanced Fields (Difficulty, Cost, Access, Region, Goals)
                difficulty_idx = get_col_index(headers, ['difficulty', 'level', 'difficulty_level'])
                cost_idx = get_col_index(headers, ['cost', 'price', 'pricing'])
                access_idx = get_col_index(headers, ['access', 'access_level', 'permissions'])
                region_idx = get_col_index(headers, ['region', 'geo', 'location'])
                goals_idx = get_col_index(headers, ['goals', 'goal', 'use_case', 'use_cases', 'purpose'])
                format_col_idx = get_col_index(headers, ['format', 'file_type'])

                difficulty = row[difficulty_idx - 1].value if difficulty_idx and len(row) >= difficulty_idx else "Intermediate"
                cost = row[cost_idx - 1].value if cost_idx and len(row) >= cost_idx else "Free"
                access = row[access_idx - 1].value if access_idx and len(row) >= access_idx else "Full Access"
                region = row[region_idx - 1].value if region_idx and len(row) >= region_idx else "Global"
                goals = row[goals_idx - 1].value if goals_idx and len(row) >= goals_idx else ""
                
                # Derive Format from Title (e.g. [Sheet], [Doc]) OR Header
                # First check column
                format_tag = "Resource"
                if format_col_idx and len(row) >= format_col_idx:
                    val = row[format_col_idx - 1].value
                    if val: format_tag = str(val).strip()

                # Fallback to brackets in title if generic
                clean_title = str(title).strip()
                if format_tag == "Resource":
                    match = re.search(r'\[(.*?)\]', clean_title)
                    if match:
                        format_tag = match.group(1).title() 

                # Create ID
                item_id = str(title).strip().lower().replace(" ", "-")
                item_id = re.sub(r'[^a-zA-Z0-9-]', '', item_id)[:60]
                
                # Clean filename for display (remove extension and parentheses)
                clean_source = os.path.basename(filepath)
                clean_source = clean_source.replace('.xlsx', '').replace('.xls', '')
                clean_source = re.sub(r' \(\d+\)', '', clean_source)

                item = {
                    "id": f"{category.lower().replace(' ', '-')}-{item_id}",
                    "category": category,
                    "title": str(title).strip(), # Keep full title for matching
                    "description": str(desc).strip(),
                    "url": str(url).strip(),
                    "type": str(item_type).strip(), # This is the "column" type
                    "format_inferred": str(format_tag).strip(), 
                    "status": str(status).strip() if status else "Active",
                    "owner": str(owner).strip() if owner else "Marketing",
                    "difficulty": str(difficulty).strip() if difficulty else "Intermediate",
                    "cost": str(cost).strip() if cost else "Free",
                    "access": str(access).strip() if access else "Full Access",
                    "region": str(region).strip() if region else "Global",
                    "goals": str(goals).strip(),
                    "tags": tags,
                    "section": current_section,
                    "source_file": clean_source
                }
                all_data.append(item)
                
        return all_data

    except Exception as e:
        print(f"Error processing {filepath}: {e}")
        return []

def main():
    all_resources = []
    
    for subdir, category_name in CATEGORIES.items():
        search_path = os.path.join(MARKETING_DIR, subdir, "*.xlsx")
        files = glob.glob(search_path)
        
        print(f"--- Processing {category_name} ({len(files)} files) ---")
        
        for f in files:
            if os.path.basename(f).startswith("~$"):
                continue
            
            print(f"Reading: {os.path.basename(f)}")
            data = process_file_with_links(f, category_name)
            all_resources.extend(data)
            print(f"  Extracted {len(data)} items")

    print(f"Total extracted items: {len(all_resources)}")
    
    # Write to JS
    js_content = f"window.MARKETING_FULL_DATA = {json.dumps(all_resources, indent=2)};"
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(js_content)
    
    print(f"Saved data to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
